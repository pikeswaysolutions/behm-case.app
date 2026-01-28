from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
import io
import csv
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'behm-funeral-home-secret-key-2024')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Security
security = HTTPBearer()

app = FastAPI(title="Behm Funeral Home Management")
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: str = "director"  # admin or director
    director_id: Optional[str] = None
    can_edit_cases: bool = False
    is_active: bool = True

class UserCreate(UserBase):
    password: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    director_id: Optional[str] = None
    can_edit_cases: Optional[bool] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None

class User(UserBase):
    id: str
    created_at: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: dict

class DirectorBase(BaseModel):
    name: str
    is_active: bool = True

class DirectorCreate(DirectorBase):
    pass

class Director(DirectorBase):
    id: str
    created_at: str

class ServiceTypeBase(BaseModel):
    name: str
    is_active: bool = True

class ServiceType(ServiceTypeBase):
    id: str

class SaleTypeBase(BaseModel):
    name: str
    is_active: bool = True

class SaleType(SaleTypeBase):
    id: str

class CaseBase(BaseModel):
    case_number: str
    date_of_death: str
    customer_first_name: str
    customer_last_name: str
    service_type_id: str
    sale_type_id: Optional[str] = None
    director_id: str
    date_paid_in_full: Optional[str] = None
    payments_received: float = 0
    average_age: Optional[float] = None
    total_sale: float = 0

class CaseCreate(CaseBase):
    pass

class CaseUpdate(BaseModel):
    case_number: Optional[str] = None
    date_of_death: Optional[str] = None
    customer_first_name: Optional[str] = None
    customer_last_name: Optional[str] = None
    service_type_id: Optional[str] = None
    sale_type_id: Optional[str] = None
    director_id: Optional[str] = None
    date_paid_in_full: Optional[str] = None
    payments_received: Optional[float] = None
    average_age: Optional[float] = None
    total_sale: Optional[float] = None

class Case(CaseBase):
    id: str
    total_balance_due: float
    created_at: str
    service_type_name: Optional[str] = None
    sale_type_name: Optional[str] = None
    director_name: Optional[str] = None

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"id": user_id, "is_active": True}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_dict = {
        "id": str(uuid.uuid4()),
        "email": user_data.email,
        "name": user_data.name,
        "role": user_data.role,
        "director_id": user_data.director_id,
        "can_edit_cases": user_data.can_edit_cases,
        "is_active": True,
        "password_hash": hash_password(user_data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_dict)
    
    token = create_access_token({"sub": user_dict["id"]})
    user_response = {k: v for k, v in user_dict.items() if k not in ["password_hash", "_id"]}
    return {"access_token": token, "token_type": "bearer", "user": user_response}

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email})
    if not user or not verify_password(credentials.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="Account is deactivated")
    
    token = create_access_token({"sub": user["id"]})
    user_response = {k: v for k, v in user.items() if k not in ["password_hash", "_id"]}
    return {"access_token": token, "token_type": "bearer", "user": user_response}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

# ==================== USER MANAGEMENT (Admin Only) ====================

@api_router.get("/users", response_model=List[User])
async def list_users(admin: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.post("/users", response_model=User)
async def create_user(user_data: UserCreate, admin: dict = Depends(require_admin)):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_dict = {
        "id": str(uuid.uuid4()),
        "email": user_data.email,
        "name": user_data.name,
        "role": user_data.role,
        "director_id": user_data.director_id,
        "can_edit_cases": user_data.can_edit_cases,
        "is_active": True,
        "password_hash": hash_password(user_data.password),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_dict)
    return {k: v for k, v in user_dict.items() if k not in ["password_hash", "_id"]}

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_update: UserUpdate, admin: dict = Depends(require_admin)):
    update_data = {k: v for k, v in user_update.model_dump().items() if v is not None}
    if "password" in update_data:
        update_data["password_hash"] = hash_password(update_data.pop("password"))
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    result = await db.users.update_one({"id": user_id}, {"$set": {"is_active": False}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deactivated"}

# ==================== FUNERAL DIRECTORS ====================

@api_router.get("/directors", response_model=List[Director])
async def list_directors(current_user: dict = Depends(get_current_user)):
    directors = await db.directors.find({}, {"_id": 0}).to_list(1000)
    return directors

@api_router.post("/directors", response_model=Director)
async def create_director(director_data: DirectorCreate, admin: dict = Depends(require_admin)):
    director_dict = {
        "id": str(uuid.uuid4()),
        "name": director_data.name,
        "is_active": director_data.is_active,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.directors.insert_one(director_dict)
    return {k: v for k, v in director_dict.items() if k != "_id"}

@api_router.put("/directors/{director_id}", response_model=Director)
async def update_director(director_id: str, director_data: DirectorCreate, admin: dict = Depends(require_admin)):
    update_data = director_data.model_dump()
    await db.directors.update_one({"id": director_id}, {"$set": update_data})
    director = await db.directors.find_one({"id": director_id}, {"_id": 0})
    if not director:
        raise HTTPException(status_code=404, detail="Director not found")
    return director

@api_router.delete("/directors/{director_id}")
async def delete_director(director_id: str, admin: dict = Depends(require_admin)):
    result = await db.directors.update_one({"id": director_id}, {"$set": {"is_active": False}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Director not found")
    return {"message": "Director deactivated"}

@api_router.post("/directors/{director_id}/reassign")
async def reassign_director_cases(
    director_id: str,
    new_director_id: str = Query(...),
    admin: dict = Depends(require_admin)
):
    result = await db.cases.update_many(
        {"director_id": director_id},
        {"$set": {"director_id": new_director_id}}
    )
    return {"message": f"Reassigned {result.modified_count} cases"}

# ==================== SERVICE TYPES ====================

@api_router.get("/service-types", response_model=List[ServiceType])
async def list_service_types(current_user: dict = Depends(get_current_user)):
    types = await db.service_types.find({}, {"_id": 0}).to_list(1000)
    return types

@api_router.post("/service-types", response_model=ServiceType)
async def create_service_type(type_data: ServiceTypeBase, admin: dict = Depends(require_admin)):
    type_dict = {
        "id": str(uuid.uuid4()),
        "name": type_data.name,
        "is_active": type_data.is_active
    }
    await db.service_types.insert_one(type_dict)
    return {k: v for k, v in type_dict.items() if k != "_id"}

@api_router.put("/service-types/{type_id}", response_model=ServiceType)
async def update_service_type(type_id: str, type_data: ServiceTypeBase, admin: dict = Depends(require_admin)):
    await db.service_types.update_one({"id": type_id}, {"$set": type_data.model_dump()})
    st = await db.service_types.find_one({"id": type_id}, {"_id": 0})
    if not st:
        raise HTTPException(status_code=404, detail="Service type not found")
    return st

@api_router.delete("/service-types/{type_id}")
async def delete_service_type(type_id: str, admin: dict = Depends(require_admin)):
    await db.service_types.update_one({"id": type_id}, {"$set": {"is_active": False}})
    return {"message": "Service type deactivated"}

# ==================== SALE TYPES ====================

@api_router.get("/sale-types", response_model=List[SaleType])
async def list_sale_types(current_user: dict = Depends(get_current_user)):
    types = await db.sale_types.find({}, {"_id": 0}).to_list(1000)
    return types

@api_router.post("/sale-types", response_model=SaleType)
async def create_sale_type(type_data: SaleTypeBase, admin: dict = Depends(require_admin)):
    type_dict = {
        "id": str(uuid.uuid4()),
        "name": type_data.name,
        "is_active": type_data.is_active
    }
    await db.sale_types.insert_one(type_dict)
    return {k: v for k, v in type_dict.items() if k != "_id"}

@api_router.put("/sale-types/{type_id}", response_model=SaleType)
async def update_sale_type(type_id: str, type_data: SaleTypeBase, admin: dict = Depends(require_admin)):
    await db.sale_types.update_one({"id": type_id}, {"$set": type_data.model_dump()})
    st = await db.sale_types.find_one({"id": type_id}, {"_id": 0})
    if not st:
        raise HTTPException(status_code=404, detail="Sale type not found")
    return st

@api_router.delete("/sale-types/{type_id}")
async def delete_sale_type(type_id: str, admin: dict = Depends(require_admin)):
    await db.sale_types.update_one({"id": type_id}, {"$set": {"is_active": False}})
    return {"message": "Sale type deactivated"}

# ==================== CASES ====================

async def enrich_case(case: dict) -> dict:
    if case.get("service_type_id"):
        st = await db.service_types.find_one({"id": case["service_type_id"]}, {"_id": 0})
        case["service_type_name"] = st["name"] if st else None
    if case.get("sale_type_id"):
        slt = await db.sale_types.find_one({"id": case["sale_type_id"]}, {"_id": 0})
        case["sale_type_name"] = slt["name"] if slt else None
    if case.get("director_id"):
        dr = await db.directors.find_one({"id": case["director_id"]}, {"_id": 0})
        case["director_name"] = dr["name"] if dr else None
    case["total_balance_due"] = case.get("total_sale", 0) - case.get("payments_received", 0)
    return case

@api_router.get("/cases", response_model=List[Case])
async def list_cases(
    current_user: dict = Depends(get_current_user),
    director_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = {}
    
    # Directors can only see their own cases
    if current_user.get("role") != "admin":
        if current_user.get("director_id"):
            query["director_id"] = current_user["director_id"]
        else:
            return []
    elif director_id:
        query["director_id"] = director_id
    
    if start_date:
        query["date_of_death"] = {"$gte": start_date}
    if end_date:
        if "date_of_death" in query:
            query["date_of_death"]["$lte"] = end_date
        else:
            query["date_of_death"] = {"$lte": end_date}
    
    cases = await db.cases.find(query, {"_id": 0}).to_list(10000)
    enriched = []
    for case in cases:
        enriched.append(await enrich_case(case))
    return enriched

@api_router.get("/cases/{case_id}", response_model=Case)
async def get_case(case_id: str, current_user: dict = Depends(get_current_user)):
    case = await db.cases.find_one({"id": case_id}, {"_id": 0})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Directors can only see their own cases
    if current_user.get("role") != "admin" and case.get("director_id") != current_user.get("director_id"):
        raise HTTPException(status_code=403, detail="Access denied")
    
    return await enrich_case(case)

@api_router.post("/cases", response_model=Case)
async def create_case(case_data: CaseCreate, current_user: dict = Depends(get_current_user)):
    # Check if director has permission to create cases
    if current_user.get("role") != "admin" and not current_user.get("can_edit_cases"):
        raise HTTPException(status_code=403, detail="You don't have permission to create cases")
    
    # Check for duplicate case number
    existing = await db.cases.find_one({"case_number": case_data.case_number})
    if existing:
        raise HTTPException(status_code=400, detail="Case number already exists")
    
    case_dict = case_data.model_dump()
    case_dict["id"] = str(uuid.uuid4())
    case_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.cases.insert_one(case_dict)
    return await enrich_case({k: v for k, v in case_dict.items() if k != "_id"})

@api_router.put("/cases/{case_id}", response_model=Case)
async def update_case(case_id: str, case_update: CaseUpdate, current_user: dict = Depends(get_current_user)):
    case = await db.cases.find_one({"id": case_id}, {"_id": 0})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Directors can only edit their own cases if they have permission
    if current_user.get("role") != "admin":
        if case.get("director_id") != current_user.get("director_id"):
            raise HTTPException(status_code=403, detail="Access denied")
        if not current_user.get("can_edit_cases"):
            raise HTTPException(status_code=403, detail="You don't have permission to edit cases")
    
    update_data = {k: v for k, v in case_update.model_dump().items() if v is not None}
    
    # Check for duplicate case number if being updated
    if "case_number" in update_data and update_data["case_number"] != case.get("case_number"):
        existing = await db.cases.find_one({"case_number": update_data["case_number"]})
        if existing:
            raise HTTPException(status_code=400, detail="Case number already exists")
    
    if update_data:
        await db.cases.update_one({"id": case_id}, {"$set": update_data})
    
    updated_case = await db.cases.find_one({"id": case_id}, {"_id": 0})
    return await enrich_case(updated_case)

@api_router.delete("/cases/{case_id}")
async def delete_case(case_id: str, admin: dict = Depends(require_admin)):
    result = await db.cases.delete_one({"id": case_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Case not found")
    return {"message": "Case deleted"}

# ==================== REPORTS & DASHBOARD ====================

@api_router.get("/reports/dashboard")
async def get_dashboard(
    current_user: dict = Depends(get_current_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    grouping: str = "monthly"
):
    query = {}
    if start_date:
        query["date_of_death"] = {"$gte": start_date}
    if end_date:
        if "date_of_death" in query:
            query["date_of_death"]["$lte"] = end_date
        else:
            query["date_of_death"] = {"$lte": end_date}
    
    # Directors only see their own data
    if current_user.get("role") != "admin" and current_user.get("director_id"):
        query["director_id"] = current_user["director_id"]
    
    cases = await db.cases.find(query, {"_id": 0}).to_list(10000)
    directors = await db.directors.find({"is_active": True}, {"_id": 0}).to_list(1000)
    director_map = {d["id"]: d["name"] for d in directors}
    
    # Calculate metrics by director
    director_metrics = {}
    for case in cases:
        did = case.get("director_id")
        if did not in director_metrics:
            director_metrics[did] = {
                "director_id": did,
                "director_name": director_map.get(did, "Unknown"),
                "case_count": 0,
                "total_sales": 0,
                "payments_received": 0,
                "total_balance_due": 0,
                "ages": []
            }
        dm = director_metrics[did]
        dm["case_count"] += 1
        dm["total_sales"] += case.get("total_sale", 0)
        dm["payments_received"] += case.get("payments_received", 0)
        dm["total_balance_due"] += case.get("total_sale", 0) - case.get("payments_received", 0)
        if case.get("average_age"):
            dm["ages"].append(case["average_age"])
    
    # Calculate average age
    for dm in director_metrics.values():
        dm["average_age"] = sum(dm["ages"]) / len(dm["ages"]) if dm["ages"] else 0
        del dm["ages"]
    
    # Grand totals
    grand_totals = {
        "case_count": sum(dm["case_count"] for dm in director_metrics.values()),
        "total_sales": sum(dm["total_sales"] for dm in director_metrics.values()),
        "payments_received": sum(dm["payments_received"] for dm in director_metrics.values()),
        "total_balance_due": sum(dm["total_balance_due"] for dm in director_metrics.values()),
        "average_age": sum(c.get("average_age", 0) for c in cases) / len(cases) if cases else 0
    }
    
    # Time series data for charts
    time_series = {}
    for case in cases:
        dod = case.get("date_of_death", "")[:7] if grouping == "monthly" else case.get("date_of_death", "")[:4]
        if dod not in time_series:
            time_series[dod] = {"period": dod, "cases": 0, "sales": 0, "payments": 0}
        time_series[dod]["cases"] += 1
        time_series[dod]["sales"] += case.get("total_sale", 0)
        time_series[dod]["payments"] += case.get("payments_received", 0)
    
    return {
        "director_metrics": list(director_metrics.values()),
        "grand_totals": grand_totals,
        "time_series": sorted(time_series.values(), key=lambda x: x["period"])
    }

@api_router.get("/reports/director/{director_id}")
async def get_director_report(
    director_id: str,
    current_user: dict = Depends(get_current_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    # Directors can only see their own reports
    if current_user.get("role") != "admin" and current_user.get("director_id") != director_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"director_id": director_id}
    if start_date:
        query["date_of_death"] = {"$gte": start_date}
    if end_date:
        if "date_of_death" in query:
            query["date_of_death"]["$lte"] = end_date
        else:
            query["date_of_death"] = {"$lte": end_date}
    
    cases = await db.cases.find(query, {"_id": 0}).to_list(10000)
    enriched_cases = [await enrich_case(c) for c in cases]
    
    metrics = {
        "case_count": len(cases),
        "total_sales": sum(c.get("total_sale", 0) for c in cases),
        "payments_received": sum(c.get("payments_received", 0) for c in cases),
        "total_balance_due": sum(c.get("total_sale", 0) - c.get("payments_received", 0) for c in cases),
        "average_age": sum(c.get("average_age", 0) for c in cases if c.get("average_age")) / max(len([c for c in cases if c.get("average_age")]), 1)
    }
    
    return {"cases": enriched_cases, "metrics": metrics}

# ==================== EXPORT ====================

@api_router.get("/export/csv")
async def export_csv(
    current_user: dict = Depends(get_current_user),
    director_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = {}
    if current_user.get("role") != "admin":
        if current_user.get("director_id"):
            query["director_id"] = current_user["director_id"]
        else:
            raise HTTPException(status_code=403, detail="No director assigned")
    elif director_id:
        query["director_id"] = director_id
    
    if start_date:
        query["date_of_death"] = {"$gte": start_date}
    if end_date:
        if "date_of_death" in query:
            query["date_of_death"]["$lte"] = end_date
        else:
            query["date_of_death"] = {"$lte": end_date}
    
    cases = await db.cases.find(query, {"_id": 0}).to_list(10000)
    enriched = [await enrich_case(c) for c in cases]
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Case Number", "Date of Death", "First Name", "Last Name", 
        "Service Type", "Sale Type", "Director", "Date Paid In Full",
        "Payments Received", "Average Age", "Total Sale", "Balance Due"
    ])
    
    for c in enriched:
        writer.writerow([
            c.get("case_number", ""),
            c.get("date_of_death", ""),
            c.get("customer_first_name", ""),
            c.get("customer_last_name", ""),
            c.get("service_type_name", ""),
            c.get("sale_type_name", ""),
            c.get("director_name", ""),
            c.get("date_paid_in_full", ""),
            c.get("payments_received", 0),
            c.get("average_age", ""),
            c.get("total_sale", 0),
            c.get("total_balance_due", 0)
        ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=cases_export.csv"}
    )

@api_router.get("/export/pdf")
async def export_pdf(
    current_user: dict = Depends(get_current_user),
    director_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = {}
    if current_user.get("role") != "admin":
        if current_user.get("director_id"):
            query["director_id"] = current_user["director_id"]
        else:
            raise HTTPException(status_code=403, detail="No director assigned")
    elif director_id:
        query["director_id"] = director_id
    
    if start_date:
        query["date_of_death"] = {"$gte": start_date}
    if end_date:
        if "date_of_death" in query:
            query["date_of_death"]["$lte"] = end_date
        else:
            query["date_of_death"] = {"$lte": end_date}
    
    cases = await db.cases.find(query, {"_id": 0}).to_list(10000)
    enriched = [await enrich_case(c) for c in cases]
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1)
    elements.append(Paragraph("Behm Funeral Home - Cases Report", title_style))
    elements.append(Spacer(1, 20))
    
    data = [["Case #", "Date of Death", "Name", "Service", "Director", "Total Sale", "Payments", "Balance"]]
    for c in enriched:
        data.append([
            c.get("case_number", ""),
            c.get("date_of_death", ""),
            f"{c.get('customer_first_name', '')} {c.get('customer_last_name', '')}",
            c.get("service_type_name", ""),
            c.get("director_name", ""),
            f"${c.get('total_sale', 0):,.2f}",
            f"${c.get('payments_received', 0):,.2f}",
            f"${c.get('total_balance_due', 0):,.2f}"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.11, 0.16, 0.25)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.97, 0.98)),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.Color(0.8, 0.85, 0.88)),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.97, 0.98)])
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=cases_report.pdf"}
    )

# ==================== IMPORT ====================

@api_router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")
    
    contents = await file.read()
    wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
    ws = wb.active
    
    # Get all directors and service/sale types for mapping
    directors = {d["name"].lower(): d["id"] for d in await db.directors.find({}, {"_id": 0}).to_list(1000)}
    service_types = {s["name"].lower(): s["id"] for s in await db.service_types.find({}, {"_id": 0}).to_list(1000)}
    sale_types = {s["name"].lower(): s["id"] for s in await db.sale_types.find({}, {"_id": 0}).to_list(1000)}
    
    imported = 0
    skipped = 0
    errors = []
    
    # Find header row
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_val = str(ws.cell(row=1, column=col).value or "").lower().strip()
        headers[cell_val] = col
    
    for row in range(2, ws.max_row + 1):
        try:
            case_num_col = headers.get("case number", headers.get("case #", 1))
            case_number = str(ws.cell(row=row, column=case_num_col).value or "").strip()
            
            if not case_number:
                continue
            
            # Check for duplicate
            existing = await db.cases.find_one({"case_number": case_number})
            if existing:
                skipped += 1
                continue
            
            # Parse data
            dod_col = headers.get("date of death", 2)
            dod_val = ws.cell(row=row, column=dod_col).value
            if isinstance(dod_val, datetime):
                date_of_death = dod_val.strftime("%Y-%m-%d")
            else:
                date_of_death = str(dod_val) if dod_val else ""
            
            first_name = str(ws.cell(row=row, column=headers.get("customer first name", headers.get("first name", 3))).value or "")
            last_name = str(ws.cell(row=row, column=headers.get("customer last name", headers.get("last name", 4))).value or "")
            
            service_type_val = str(ws.cell(row=row, column=headers.get("service type", 5)).value or "").lower().strip()
            service_type_id = service_types.get(service_type_val)
            if not service_type_id and service_type_val:
                # Create new service type
                new_st = {"id": str(uuid.uuid4()), "name": service_type_val.title(), "is_active": True}
                await db.service_types.insert_one(new_st)
                service_types[service_type_val] = new_st["id"]
                service_type_id = new_st["id"]
            
            sale_type_val = str(ws.cell(row=row, column=headers.get("sale type", 6)).value or "").lower().strip()
            sale_type_id = sale_types.get(sale_type_val)
            if not sale_type_id and sale_type_val:
                new_slt = {"id": str(uuid.uuid4()), "name": sale_type_val.title(), "is_active": True}
                await db.sale_types.insert_one(new_slt)
                sale_types[sale_type_val] = new_slt["id"]
                sale_type_id = new_slt["id"]
            
            director_val = str(ws.cell(row=row, column=headers.get("director", headers.get("funeral director", 7))).value or "").lower().strip()
            director_id = directors.get(director_val)
            if not director_id and director_val:
                new_dir = {"id": str(uuid.uuid4()), "name": director_val.title(), "is_active": True, "created_at": datetime.now(timezone.utc).isoformat()}
                await db.directors.insert_one(new_dir)
                directors[director_val] = new_dir["id"]
                director_id = new_dir["id"]
            
            def parse_float(val):
                if val is None:
                    return 0
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return 0
            
            def parse_date(val):
                if val is None:
                    return None
                if isinstance(val, datetime):
                    return val.strftime("%Y-%m-%d")
                return str(val) if val else None
            
            case_dict = {
                "id": str(uuid.uuid4()),
                "case_number": case_number,
                "date_of_death": date_of_death,
                "customer_first_name": first_name,
                "customer_last_name": last_name,
                "service_type_id": service_type_id or "",
                "sale_type_id": sale_type_id,
                "director_id": director_id or "",
                "date_paid_in_full": parse_date(ws.cell(row=row, column=headers.get("date paid in full", headers.get("pif", 8))).value),
                "payments_received": parse_float(ws.cell(row=row, column=headers.get("payments received", 9)).value),
                "average_age": parse_float(ws.cell(row=row, column=headers.get("avg age", headers.get("average age", 10))).value),
                "total_sale": parse_float(ws.cell(row=row, column=headers.get("total sale", 11)).value),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.cases.insert_one(case_dict)
            imported += 1
            
        except Exception as e:
            errors.append(f"Row {row}: {str(e)}")
    
    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:10]  # Limit errors returned
    }

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data():
    # Check if already seeded
    existing_admin = await db.users.find_one({"email": "admin@behmfuneral.com"})
    if existing_admin:
        return {"message": "Data already seeded"}
    
    # Create default service types
    service_types = [
        {"id": str(uuid.uuid4()), "name": "Traditional", "is_active": True},
        {"id": str(uuid.uuid4()), "name": "Cremation", "is_active": True},
        {"id": str(uuid.uuid4()), "name": "Memorial", "is_active": True},
        {"id": str(uuid.uuid4()), "name": "Graveside", "is_active": True}
    ]
    await db.service_types.insert_many(service_types)
    
    # Create default sale types
    sale_types = [
        {"id": str(uuid.uuid4()), "name": "At-Need", "is_active": True},
        {"id": str(uuid.uuid4()), "name": "Pre-Need", "is_active": True},
        {"id": str(uuid.uuid4()), "name": "Insurance", "is_active": True}
    ]
    await db.sale_types.insert_many(sale_types)
    
    # Create sample directors
    directors = [
        {"id": str(uuid.uuid4()), "name": "Eric Behm", "is_active": True, "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "name": "Trevor Behm", "is_active": True, "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "name": "John Merk", "is_active": True, "created_at": datetime.now(timezone.utc).isoformat()}
    ]
    await db.directors.insert_many(directors)
    
    # Create admin user
    admin_user = {
        "id": str(uuid.uuid4()),
        "email": "admin@behmfuneral.com",
        "name": "Administrator",
        "role": "admin",
        "director_id": None,
        "can_edit_cases": True,
        "is_active": True,
        "password_hash": hash_password("admin123"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_user)
    
    # Create director user
    director_user = {
        "id": str(uuid.uuid4()),
        "email": "eric@behmfuneral.com",
        "name": "Eric Behm",
        "role": "director",
        "director_id": directors[0]["id"],
        "can_edit_cases": True,
        "is_active": True,
        "password_hash": hash_password("director123"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(director_user)
    
    # Create sample cases
    import random
    for i in range(25):
        director = random.choice(directors)
        service_type = random.choice(service_types)
        sale_type = random.choice(sale_types)
        total_sale = random.uniform(3000, 15000)
        payments = random.uniform(0, total_sale)
        
        case = {
            "id": str(uuid.uuid4()),
            "case_number": f"BFH-2024-{str(i+1).zfill(4)}",
            "date_of_death": f"2024-{str(random.randint(1, 12)).zfill(2)}-{str(random.randint(1, 28)).zfill(2)}",
            "customer_first_name": random.choice(["John", "Mary", "Robert", "Patricia", "William", "Linda", "James", "Barbara"]),
            "customer_last_name": random.choice(["Smith", "Johnson", "Williams", "Brown", "Jones", "Miller", "Davis", "Garcia"]),
            "service_type_id": service_type["id"],
            "sale_type_id": sale_type["id"],
            "director_id": director["id"],
            "date_paid_in_full": f"2024-{str(random.randint(1, 12)).zfill(2)}-{str(random.randint(1, 28)).zfill(2)}" if payments >= total_sale else None,
            "payments_received": round(payments, 2),
            "average_age": random.randint(55, 95),
            "total_sale": round(total_sale, 2),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.cases.insert_one(case)
    
    return {"message": "Data seeded successfully"}

@api_router.get("/")
async def root():
    return {"message": "Behm Funeral Home Management API"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
